#功能：测试excel
#说明: Python版本 3.13.5

import openpyxl

#读excel文件并按分数排序
def read_excel(filepath) -> None :
    # 打开Excel文件
    workbook = openpyxl.load_workbook(filepath)

    #获取当前活动工作表（默认为第一张）
    sheet = workbook.active
 
    # 存储学生数据的列表
    students = []
    
    #遍历每一行，读取学生ID和分数（跳过表头）
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2, values_only=True):
        if row[0] is not None and row[1] is not None:  # 确保数据不为空
            student_id = row[0]
            score = row[1]
            students.append((student_id, score))
    
    # 按分数排序（从高到低）
    students.sort(key=lambda x: x[1], reverse=True)
    
    # 输出排序后的结果
    print(f"{'序号':<4} {'学生ID':<10} {'分数':<6}")
    print("-" * 22)
    for index, (student_id, score) in enumerate(students, 1):
        print(f"{index:<4} {student_id:<10} {score:<6}")
    
    return 

def main():
    filepath = "file//CSX2006_M1_Sec544.xlsx"
    read_excel(filepath)
    return 


if __name__ == "__main__":
    main()