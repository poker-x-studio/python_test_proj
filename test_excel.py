#功能：测试excel
#说明: Python版本 3.13.5

import openpyxl

#读excel文件
def read_excel(filepath) -> None :
    # 打开Excel文件
    workbook = openpyxl.load_workbook(filepath)

    #获取当前活动工作表（默认为第一张）
    sheet = workbook.active

    
    #遍历每一行并输出A列的值
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=1, values_only=True):
        for cell in row:
            print(cell) 

    sum = sheet['A1'].value + sheet['A2'].value   
    print(sum)    

    #写入
    sheet["A3"] = sum
    workbook.save(filepath)
    return 

#写excel文件
def write_excel(filepath) -> None :
    # 创建新的Excel文件并添加数据
    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active
    new_sheet["A1"] = 30
    new_sheet["A2"] = 40
    new_workbook.save(filepath)
    return 


def main():
    filepath = "file//test.xlsx"
    #write_excel(filepath)
    read_excel(filepath)
    return 


if __name__ == "__main__":
    main()